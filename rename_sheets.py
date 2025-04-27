import os
from openpyxl import load_workbook
from openpyxl.utils.exceptions import IllegalCharacterError

def rename_sheets_based_on_c5(file_path):
    """
    Renames all sheets in an Excel file based on the value in cell C5 of each sheet.
    
    Args:
        file_path (str): Path to the Excel file
    
    Returns:
        tuple: (success: bool, message: str)
    """
    try:
        # Load the workbook
        wb = load_workbook(filename=file_path)
        changes_made = False
        
        # Iterate through each sheet
        for sheet in wb.sheetnames:
            current_sheet = wb[sheet]
            
            # Get the value from cell C5
            nn=current_sheet['C5'].value
            print(nn)
            new_name = current_sheet['C5'].value 
            print(new_name)
            
            # Check if C5 has a value and it's different from current name
            if new_name and new_name != sheet:
                try:
                    # Rename the sheet
                    current_sheet.title = str(new_name)[:31]  # Truncate to 31 chars max
                    changes_made = True
                    print(f"  Renamed sheet '{sheet}' to '{new_name}'")
                except IllegalCharacterError:
                    print(f"  Could not rename sheet '{sheet}' - invalid characters in name")
                except Exception as e:
                    print(f"  Could not rename sheet '{sheet}': {str(e)}")
        
        # Save the workbook if changes were made
        if changes_made:
            wb.save(file_path)
            return (True, f"Updated: {os.path.basename(file_path)}")
        return (True, f"No changes needed: {os.path.basename(file_path)}")
    
    except Exception as e:
        return (False, f"Error processing {os.path.basename(file_path)}: {str(e)}")

def process_excel_files_in_folder(folder_path):
    """
    Processes all Excel files in a folder, renaming sheets based on cell C5.
    
    Args:
        folder_path (str): Path to the folder containing Excel files
    """
    if not os.path.exists(folder_path):
        print(f"Error: Folder not found - {folder_path}")
        return
    
    print(f"\nProcessing Excel files in: {folder_path}")
    
    # Supported Excel extensions
    excel_extensions = ('.xlsx', '.xlsm', '.xltx', '.xltm')
    
    # Process each file in folder
    processed_files = 0
    for filename in os.listdir(folder_path):
        if filename.lower().endswith(excel_extensions):
            file_path = os.path.join(folder_path, filename)
            print(f"\nProcessing file: {filename}")
            
            success, message = rename_sheets_based_on_c5(file_path)
            print(f"  {message}")
            
            if success:
                processed_files += 1
    
    print(f"\nProcessing complete. {processed_files} files were processed.")

# Example usage
if __name__ == "__main__":
    folder_path = input("Enter the path to the folder containing Excel files: ").strip()
    process_excel_files_in_folder(folder_path)
