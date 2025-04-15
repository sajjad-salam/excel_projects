import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def count_voters(input_file):
    # Extract the file name without extension
    file_name = os.path.basename(input_file)
    file_name_without_ext = os.path.splitext(file_name)[0]

    # Load the workbook
    try:
        wb = load_workbook(input_file)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    # Create or clear Results and غير المحدثين sheets
    if "الاحصائية" in wb.sheetnames:
        results_sheet = wb["الاحصائية"]
        results_sheet.delete_rows(1, results_sheet.max_row)
    else:
        results_sheet = wb.create_sheet("الاحصائية")

    if "غير المحدثين" in wb.sheetnames:
        update_sheet = wb["غير المحدثين"]
        update_sheet.delete_rows(1, update_sheet.max_row)
    else:
        update_sheet = wb.create_sheet("غير المحدثين")

    # Add title to الاحصائية sheet
    title_font = Font(size=24, bold=True)  # Larger font for title
    center_alignment = Alignment(horizontal="center", vertical="center")
    results_sheet.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=4)
    title_cell = results_sheet.cell(
        row=1, column=1, value=f"الاحصائية للملف : {file_name_without_ext}")
    title_cell.font = title_font
    title_cell.alignment = center_alignment

    # Set up headers in Results sheet (start from row 3)
    # change font size to 20
    results_sheet.append(["اسم الورقة", "عدد التحديث",
                         "عدد تم التحديث", "عدد البطايق"])
    header_fill = PatternFill(start_color="BFBFBF",
                              end_color="BFBFBF", fill_type="solid")
    bold_font = Font(bold=True, size=20)  # Set font size to 20
    # results_sheet.append(["اسم الورقة", "عدد التحديث",
    #                      "عدد تم التحديث", "عدد البطايق"])

    results_output_row = 3  # Start data from row 3

    # Add title to غير المحدثين sheet
    update_sheet.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=3)
    title_cell = update_sheet.cell(
        row=1, column=1, value=f"الاسماء الغير محدثة للملف : {file_name_without_ext}")
    title_cell.font = title_font
    title_cell.alignment = center_alignment

    # Set up headers in غير المحدثين sheet (start from row 3)
    update_sheet.append(["ت", "اسم الناخب الثلاثي", "المسؤول عنه"])
    header_fill = PatternFill(start_color="BFBFBF",
                              end_color="BFBFBF", fill_type="solid")
    bold_font = Font(bold=True, size=20)  # Set font size to 20
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for col in range(1, 4):
        cell = update_sheet.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border

    update_output_row = 4  # Start data from row 4

    # Loop through all sheets (excluding Results and غير المحدثين)
    for sheet_name in wb.sheetnames:
        if sheet_name not in ["الاحصائية", "غير المحدثين"]:
            ws = wb[sheet_name]

            count_tahdeeth = 0
            count_tam_tahdeeth = 0
            total_people = 0

            # Remove "10_ورقة1" from sheet name for responsible person
            responsible_name = sheet_name.replace("10_ورقة1", "")

            # Find last row in column C (المركز الانتخابي)
            last_row = ws.max_row
            if last_row < 9:
                last_row = 9  # Ensure at least row 9 is checked
            if last_row > 100:
                last_row = 100  # Don't go beyond row 100

            # Loop through data rows (9 to lastRow)
            for i in range(9, last_row + 1):
                if ws.cell(row=i, column=3).value is not None:  # Check column C
                    total_people += 1
                    cell_value = str(ws.cell(row=i, column=3).value).strip()

                    # Check for تم التحديث (exact match first)
                    if cell_value == "تم التحديث":
                        count_tam_tahdeeth += 1
                    # Check for تحديث (exact match)
                    elif cell_value == "تحديث":
                        count_tahdeeth += 1
                        # Add to غير المحدثين sheet
                        update_sheet.append([update_output_row - 1,
                                             ws.cell(row=i, column=2).value,
                                             responsible_name])
                        update_output_row += 1
                    # Check for partial matches if needed
                    elif "تم التحديث" in cell_value:
                        count_tam_tahdeeth += 1
                    elif "تحديث" in cell_value:
                        count_tahdeeth += 1
                        # Add to غير المحدثين sheet
                        update_sheet.append([update_output_row - 1,
                                             ws.cell(row=i, column=2).value,
                                             responsible_name])
                        update_output_row += 1

            # Write results to Results sheet
            results_sheet.append([sheet_name, count_tahdeeth, count_tam_tahdeeth,
                                  total_people - count_tahdeeth - count_tam_tahdeeth])
            results_output_row += 1

    # Format Results sheet
    results_sheet.auto_filter.ref = f"A3:D{results_output_row - 1}"
    for row in results_sheet.iter_rows(min_row=3, max_row=results_output_row - 1, min_col=1, max_col=4):
        for cell in row:
            cell.font = Font(size=20)  # Set font size to 20
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # Adjust column widths to fit text
    for col in range(1, 5):  # Columns A to D
        max_length = 0
        column = get_column_letter(col)
        for cell in results_sheet[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Add padding and scale
        results_sheet.column_dimensions[column].width = adjusted_width

    # Add totals row
    if results_output_row > 3:
        results_sheet.append(["المجموع", f"=SUM(B3:B{results_output_row - 1})",
                              f"=SUM(C3:C{results_output_row - 1})",
                              f"=SUM(D3:D{results_output_row - 1})"])
        for cell in results_sheet[results_output_row]:
            cell.font = Font(bold=True, size=20)  # Set font size to 20

    # Format غير المحدثين sheet
    if update_output_row > 4:
        for row in update_sheet.iter_rows(min_row=4, max_row=update_output_row - 1, min_col=1, max_col=3):
            for cell in row:
                cell.font = Font(size=20)  # Set font size to 20
                cell.alignment = Alignment(
                    horizontal="center", vertical="center")
                cell.border = thin_border

        # Adjust column widths to fit text
        for col in range(1, 4):  # Columns A to C
            max_length = 0
            column = get_column_letter(col)
            for cell in update_sheet[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2  # Add padding and scale
            update_sheet.column_dimensions[column].width = adjusted_width
    else:
        update_sheet.append(["", "لا يوجد ناخبين غير محدثين", ""])

    # Set sheets to right-to-left
    results_sheet.sheet_view.rightToLeft = True
    update_sheet.sheet_view.rightToLeft = True

    # Save the workbook
    wb.save(input_file)
    print("تم الانتهاء من عملية العد بنجاح!")
    print("النتائج موجودة في ورقة 'الاحصائية'")
    print("قائمة غير المحدثين موجودة في ورقة 'غير المحدثين'")


# تحتاج الى ملف بي شيتات ابو ال10 بداخلة
# راح يرجعلك نفس الملف بس بداخلة شيت الاحصائية وشيت الغير محدثين
# Run the function
input_file = "copied_data.xlsx"
count_voters(input_file)
