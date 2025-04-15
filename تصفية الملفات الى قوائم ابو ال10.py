import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from openpyxl.styles import NamedStyle
from copy import copy
import shutil
original_file = 'input.xlsx'
new_file = input("Enter your file name (e.g., 'new_file.xlsx'): ")
new_file = f"{new_file}.xlsx"

copied_file = 'copied_' + new_file
shutil.copyfile(original_file, copied_file)  # Copy the file


def transfer_all_data_with_format():
    # Load source data
    data_wb = openpyxl.load_workbook(new_file)
    data_ws = data_wb.active

    # Convert to DataFrame
    data = []
    for row in data_ws.iter_rows(min_row=3, max_row=data_ws.max_row, min_col=1, max_col=7, values_only=True):
        data.append(row)

    df = pd.DataFrame(data, columns=[
                      'ت', 'الاسم', 'المهنة', 'التحديث', 'المركز الانتخابي', 'السكن', 'عمود1'])

    # Get all unique values in column G (عمود1)
    unique_values = df['عمود1'].unique()

    # Load target workbook
    input_wb = openpyxl.load_workbook(copied_file)

    # Get the template sheet (Sheet1)
    if 'Sheet1' not in input_wb.sheetnames:
        raise Exception("Template sheet 'Sheet1' not found in input.xlsx")

    template_sheet = input_wb['Sheet1']

    # Process each unique value
    # Process each unique value
    for value in unique_values:
        if pd.isna(value):  # Skip NaN values
            continue

        # Skip if sheet name matches column G value (عمود1)
        if str(value).strip() == 'عمود1':
            continue

        # Filter data for this value
        filtered = df[df['عمود1'] == value]

        # Create sheet name (replace invalid characters)
        sheet_name = str(value)[:30]  # Excel sheet name limit

        # Create new sheet by copying template or use existing
        if sheet_name not in input_wb.sheetnames:
            # Copy the template sheet
            input_ws = input_wb.copy_worksheet(template_sheet)
            input_ws.title = sheet_name
        else:
            input_ws = input_wb[sheet_name]

        # Rest of your existing code...
        # Set sheet to right-to-left
        input_ws.sheet_view.rightToLeft = True

        # Add sheet name in cell C5
        input_ws['C5'] = value

        # Clear existing data (from row 9 to 22) but keep formatting
        for row in input_ws.iter_rows(min_row=9, max_row=22, min_col=1, max_col=3):
            for cell in row:
                cell.value = None

        # Write the filtered data
        name_count = 0  # Initialize counter for names
        for i, (index, row) in enumerate(filtered.iterrows(), start=9):
            if i > 22:  # Don't exceed row 22
                break

            # Write to columns B and C
            input_ws.cell(row=i, column=2, value=row['الاسم'])
            input_ws.cell(row=i, column=3, value=row['المركز الانتخابي'])

            # Write the index number in column A
            input_ws.cell(row=i, column=1, value=i-8)

            name_count += 1  # Increment name counter

        # Add name count in C25 with label
        # حساب عدد الاسماء مال الناخبين واضافته الى الشيت
        # اني لغيته لان سويته تلقائي ب اكسل احسن حتى يتحدث اوتماتيك
        # input_ws['G10'] = name_count

    # delete sheet1
    if 'Sheet1' in input_wb.sheetnames:
        std = input_wb['Sheet1']
        input_wb.remove(std)
    # Save the changes
    input_wb.save(copied_file)
    input_wb.close()
    #
    print(
        f"Data transferred successfully to {len(unique_values)} sheets with consistent formatting!")


# Run the automated transfer
transfer_all_data_with_format()
